"""
export_utils.py — Utilidades de exportación PDF y Excel reutilizables
Usadas por InstApp y AdminApp.
"""
import io
from django.http import HttpResponse

# ──────────────────────────────────────────────
#  PDF (reportlab)
# ──────────────────────────────────────────────
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT


def generar_pdf_response(filename: str) -> tuple:
    """
    Devuelve (response, buffer) listos para usar con SimpleDocTemplate.
    Uso:
        response, buffer = generar_pdf_response("mi_reporte.pdf")
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
        ...construir elements...
        doc.build(elements)
        response.write(buffer.getvalue())
        return response
    """
    buffer = io.BytesIO()
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response, buffer


def _estilos_tabla():
    return TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2d6a4f")),
        ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",   (0, 0), (-1, 0), 10),
        ("ALIGN",      (0, 0), (-1, -1), "CENTER"),
        ("VALIGN",     (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f0f9f4")]),
        ("GRID",       (0, 0), (-1, -1), 0.5, colors.HexColor("#cccccc")),
        ("FONTSIZE",   (0, 1), (-1, -1), 9),
        ("TOPPADDING",  (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("LEFTPADDING",   (0, 0), (-1, -1), 8),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 8),
    ])


def construir_pdf(
    buffer,
    titulo: str,
    cabeceras: list,
    filas: list,
    orientacion="horizontal",
):
    """
    Construye un PDF limpio tipo reporte en el buffer dado.
    cabeceras: lista de strings
    filas: lista de listas de strings
    orientacion: 'horizontal' o 'vertical'
    """
    pagesize = landscape(A4) if orientacion == "horizontal" else A4
    doc = SimpleDocTemplate(
        buffer,
        pagesize=pagesize,
        rightMargin=1.5 * cm,
        leftMargin=1.5 * cm,
        topMargin=1.5 * cm,
        bottomMargin=1.5 * cm,
    )
    styles = getSampleStyleSheet()
    titulo_style = ParagraphStyle(
        "titulo",
        parent=styles["Heading1"],
        fontSize=14,
        textColor=colors.HexColor("#2d6a4f"),
        alignment=TA_CENTER,
        spaceAfter=12,
    )
    elements = [
        Paragraph(titulo, titulo_style),
        Spacer(1, 0.3 * cm),
    ]

    texto_style = ParagraphStyle(
        "texto_tabla",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=8,
        leading=10,
        alignment=TA_LEFT,
    )
    header_style = ParagraphStyle(
        "header_tabla",
        parent=styles["BodyText"],
        fontName="Helvetica-Bold",
        fontSize=9,
        leading=11,
        alignment=TA_CENTER,
        textColor=colors.white,
    )

    def _cell_len(value):
        return max(4, min(len(str(value or "")), 40))

    wrapped_headers = [Paragraph(str(c), header_style) for c in cabeceras]
    wrapped_rows = [
        [Paragraph(str(valor if valor is not None else ""), texto_style) for valor in fila]
        for fila in filas
    ]
    datos = [wrapped_headers] + wrapped_rows

    total_width = doc.width
    col_weights = []
    for col_idx, cabecera in enumerate(cabeceras):
        max_len = _cell_len(cabecera)
        for fila in filas:
            if col_idx < len(fila):
                max_len = max(max_len, _cell_len(fila[col_idx]))
        col_weights.append(max_len)

    weight_sum = sum(col_weights) or len(cabeceras)
    col_widths = [(peso / weight_sum) * total_width for peso in col_weights]
    tabla = Table(datos, repeatRows=1, colWidths=col_widths)
    tabla.setStyle(_estilos_tabla())
    elements.append(tabla)
    doc.build(elements)


# ──────────────────────────────────────────────
#  EXCEL (openpyxl)
# ──────────────────────────────────────────────
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def generar_excel_response(filename: str) -> tuple:
    """
    Devuelve (response, workbook, worksheet).
    Llamar response.write(...) al final.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response, wb, ws


def estilizar_excel(ws, cabeceras: list, filas: list, titulo: str = ""):
    """
    Escribe cabeceras + filas en ws con estilo institucional verde.
    """
    VERDE = "2D6A4F"
    VERDE_CLARO = "D8F3DC"

    # Título en fila 1 (opcional)
    fila_inicio = 1
    if titulo:
        ws.merge_cells(f"A1:{get_column_letter(len(cabeceras))}1")
        cell_titulo = ws.cell(row=1, column=1, value=titulo)
        cell_titulo.font = Font(bold=True, color="FFFFFF", size=13)
        cell_titulo.fill = PatternFill("solid", fgColor=VERDE)
        cell_titulo.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 22
        fila_inicio = 2

    # Cabeceras
    for col_idx, cab in enumerate(cabeceras, start=1):
        cell = ws.cell(row=fila_inicio, column=col_idx, value=cab)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor=VERDE)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[fila_inicio].height = 18

    # Datos
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for fila_idx, fila in enumerate(filas, start=fila_inicio + 1):
        fill_color = VERDE_CLARO if (fila_idx - fila_inicio) % 2 == 0 else "FFFFFF"
        for col_idx, valor in enumerate(fila, start=1):
            cell = ws.cell(row=fila_idx, column=col_idx, value=str(valor) if valor is not None else "")
            cell.fill = PatternFill("solid", fgColor=fill_color)
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[fila_idx].height = 15

    # Ajuste de ancho de columnas
    for col_idx in range(1, len(cabeceras) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(cabeceras[col_idx - 1]))
        for fila in filas:
            if col_idx - 1 < len(fila):
                cell_val = str(fila[col_idx - 1]) if fila[col_idx - 1] else ""
                max_len = max(max_len, len(cell_val))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)


def guardar_excel_en_response(response, wb):
    buffer = io.BytesIO()
    wb.save(buffer)
    response.write(buffer.getvalue())
    return response
