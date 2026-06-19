from io import BytesIO

from django.http import HttpResponse
from django.views.decorators.cache import never_cache

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from LoginApp.models import Ambiente, RegistroIncidente, RegistroMinuta, TrasladoRecurso

from .utils import (
    _no_cache,
    _filters_to_suffix,
    _get_ambientes_filters,
    _aplicar_filtros_ambientes,
    _get_incidentes_filters,
    _aplicar_filtros_incidentes,
    _get_minutas_filters,
    _aplicar_filtros_minutas,
    _get_traslados_filters,
    _aplicar_filtros_traslados,
)


def _estilizar_excel_guarda(ws):
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 45)

    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 30


def _pdf_table_styles(colors):
    from reportlab.platypus import TableStyle

    return TableStyle(
        [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f766e")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.whitesmoke]),
            ("LEFTPADDING", (0, 0), (-1, -1), 5),
            ("RIGHTPADDING", (0, 0), (-1, -1), 5),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]
    )


def _build_wrapped_pdf_data(headers, rows, styles, colors):
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.platypus import Paragraph

    header_style = ParagraphStyle(
        "guarda_pdf_header",
        parent=styles["BodyText"],
        fontName="Helvetica-Bold",
        fontSize=8,
        leading=10,
        alignment=TA_CENTER,
        textColor=colors.white,
    )
    cell_style = ParagraphStyle(
        "guarda_pdf_cell",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=7,
        leading=9,
        alignment=TA_LEFT,
    )

    data = [[Paragraph(str(header), header_style) for header in headers]]
    for row in rows:
        data.append([
            Paragraph(str(value if value is not None else ""), cell_style)
            for value in row
        ])
    return data


def _pdf_doc(title, filtros, landscape_mode=True):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer

    styles = getSampleStyleSheet()
    buf = BytesIO()
    pagesize = landscape(A4) if landscape_mode else A4
    doc = SimpleDocTemplate(
        buf,
        pagesize=pagesize,
        leftMargin=1.0 * cm,
        rightMargin=1.0 * cm,
        topMargin=1.2 * cm,
        bottomMargin=1.2 * cm,
    )
    elements = [Paragraph(title, styles["Title"]), Spacer(1, 12)]
    subtitle = f"Filtro: {_filters_to_suffix(filtros)}" if any(filtros.values()) else "Sin filtro"
    elements.append(Paragraph(subtitle, styles["BodyText"]))
    elements.append(Spacer(1, 12))
    return buf, doc, styles, elements


def _excel_response(filename, ws):
    buf = BytesIO()
    ws.parent.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return _no_cache(resp)


@never_cache
def exportar_minutas_excel(request):
    filtros = _get_minutas_filters(request)
    minutas = (
        RegistroMinuta.objects.select_related(
            "ambiente",
            "guarda_seguridad_usuario_id_usuario__usuario_id_usuario",
            "responsable__usuario_id_usuario",
        )
        .all()
        .order_by("-fecha_hora_recibo")
    )
    minutas = _aplicar_filtros_minutas(minutas, filtros)

    wb = Workbook()
    ws = wb.active
    ws.title = "Minutas"
    headers = ["ID", "Ambiente", "Recibo", "Entrega", "Estado", "Novedad", "Descripcion", "Responsable", "Guarda"]
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col_idx)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for m in minutas:
        responsable = f"{getattr(m.responsable.usuario_id_usuario, 'p_nombre', '')} {getattr(m.responsable.usuario_id_usuario, 'p_apellido', '')}".strip()
        guarda = f"{getattr(m.guarda_seguridad_usuario_id_usuario.usuario_id_usuario, 'p_nombre', '')} {getattr(m.guarda_seguridad_usuario_id_usuario.usuario_id_usuario, 'p_apellido', '')}".strip()
        ws.append([
            m.id_minuta,
            m.ambiente.num_ambiente if m.ambiente_id else "",
            m.fecha_hora_recibo.strftime("%Y-%m-%d %H:%M:%S") if m.fecha_hora_recibo else "",
            m.fecha_hora_entrega.strftime("%Y-%m-%d %H:%M:%S") if m.fecha_hora_entrega else "",
            m.estado or "",
            m.novedad or "",
            m.descripcion_min or "",
            responsable,
            guarda,
        ])

    _estilizar_excel_guarda(ws)
    return _excel_response(f"minutas_{_filters_to_suffix(filtros)}.xlsx", ws)


@never_cache
def exportar_minutas_pdf(request):
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import Table

    filtros = _get_minutas_filters(request)
    minutas = (
        RegistroMinuta.objects.select_related(
            "ambiente",
            "guarda_seguridad_usuario_id_usuario__usuario_id_usuario",
            "responsable__usuario_id_usuario",
        )
        .all()
        .order_by("-fecha_hora_recibo")
    )
    minutas = _aplicar_filtros_minutas(minutas, filtros)

    headers = ["ID", "Ambiente", "Recibo", "Entrega", "Estado", "Novedad", "Descripcion"]
    rows = []
    for m in minutas:
        rows.append([
            str(m.id_minuta),
            str(m.ambiente.num_ambiente) if m.ambiente_id else "",
            m.fecha_hora_recibo.strftime("%Y-%m-%d %H:%M:%S") if m.fecha_hora_recibo else "",
            m.fecha_hora_entrega.strftime("%Y-%m-%d %H:%M:%S") if m.fecha_hora_entrega else "",
            m.estado or "",
            m.novedad or "",
            (m.descripcion_min or "")[:300],
        ])

    buf, doc, styles, elements = _pdf_doc("Reporte de Minutas", filtros, landscape_mode=True)
    data = _build_wrapped_pdf_data(headers, rows, styles, colors)
    table = Table(data, repeatRows=1, colWidths=[1.2 * cm, 2.0 * cm, 3.0 * cm, 3.0 * cm, 2.4 * cm, 4.0 * cm, 9.4 * cm])
    table.setStyle(_pdf_table_styles(colors))
    elements.append(table)
    doc.build(elements)

    resp = HttpResponse(buf.getvalue(), content_type="application/pdf")
    resp["Content-Disposition"] = f'attachment; filename="minutas_{_filters_to_suffix(filtros)}.pdf"'
    return _no_cache(resp)


@never_cache
def exportar_incidentes_excel(request):
    filtros = _get_incidentes_filters(request)
    incidentes = (
        RegistroIncidente.objects.select_related("ambiente", "tipo_inc", "usuario_id_usuario")
        .all()
        .order_by("-fecha_incidente", "-hora_incidente")
    )
    incidentes = _aplicar_filtros_incidentes(incidentes, filtros)

    wb = Workbook()
    ws = wb.active
    ws.title = "Incidentes"
    headers = ["ID", "Fecha", "Hora", "Ambiente", "Tipo", "Gravedad", "Descripcion", "Usuario"]
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col_idx)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for i in incidentes:
        usuario = f"{i.usuario_id_usuario.p_nombre} {i.usuario_id_usuario.p_apellido}".strip()
        ws.append([
            i.id_incidente,
            i.fecha_incidente.isoformat() if i.fecha_incidente else "",
            str(i.hora_incidente) if i.hora_incidente else "",
            i.ambiente.num_ambiente if i.ambiente_id else "",
            i.tipo_inc.tipo_incidente if i.tipo_inc_id else "",
            i.nivel_gravedad or "",
            i.descripcion or "",
            usuario,
        ])

    _estilizar_excel_guarda(ws)
    return _excel_response(f"incidentes_{_filters_to_suffix(filtros)}.xlsx", ws)


@never_cache
def exportar_incidentes_pdf(request):
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import Table

    filtros = _get_incidentes_filters(request)
    incidentes = (
        RegistroIncidente.objects.select_related("ambiente", "tipo_inc", "usuario_id_usuario")
        .all()
        .order_by("-fecha_incidente", "-hora_incidente")
    )
    incidentes = _aplicar_filtros_incidentes(incidentes, filtros)

    headers = ["ID", "Fecha", "Hora", "Ambiente", "Tipo", "Gravedad", "Descripcion", "Usuario"]
    rows = []
    for i in incidentes:
        usuario = f"{i.usuario_id_usuario.p_nombre} {i.usuario_id_usuario.p_apellido}".strip()
        rows.append([
            str(i.id_incidente),
            i.fecha_incidente.isoformat() if i.fecha_incidente else "",
            str(i.hora_incidente) if i.hora_incidente else "",
            str(i.ambiente.num_ambiente) if i.ambiente_id else "",
            i.tipo_inc.tipo_incidente if i.tipo_inc_id else "",
            i.nivel_gravedad or "",
            (i.descripcion or "")[:300],
            usuario,
        ])

    buf, doc, styles, elements = _pdf_doc("Reporte de Incidentes", filtros, landscape_mode=True)
    data = _build_wrapped_pdf_data(headers, rows, styles, colors)
    table = Table(data, repeatRows=1, colWidths=[1.0 * cm, 2.0 * cm, 1.8 * cm, 1.8 * cm, 2.8 * cm, 2.0 * cm, 7.8 * cm, 4.8 * cm])
    table.setStyle(_pdf_table_styles(colors))
    elements.append(table)
    doc.build(elements)

    resp = HttpResponse(buf.getvalue(), content_type="application/pdf")
    resp["Content-Disposition"] = f'attachment; filename="incidentes_{_filters_to_suffix(filtros)}.pdf"'
    return _no_cache(resp)


@never_cache
def exportar_traslados_excel(request):
    filtros = _get_traslados_filters(request)
    traslados = (
        TrasladoRecurso.objects.select_related("recurso", "ambiente_origen", "instructor_origen__usuario_id_usuario", "instructor_destino__usuario_id_usuario")
        .all()
        .order_by("-fecha_traslado")
    )
    traslados = _aplicar_filtros_traslados(traslados, filtros)

    destinos_ids = {t.ambiente_destino for t in traslados if t.ambiente_destino}
    destinos_map = {a.id_ambiente: a.num_ambiente for a in Ambiente.objects.filter(id_ambiente__in=destinos_ids)}

    wb = Workbook()
    ws = wb.active
    ws.title = "Traslados"
    headers = ["ID", "Recurso", "Serial", "Origen", "Destino", "Fecha", "Instructor Presta", "Instructor Recibe", "Duracion", "Observacion"]
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col_idx)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for t in traslados:
        instructor_origen = f"{t.instructor_origen.usuario_id_usuario.p_nombre} {t.instructor_origen.usuario_id_usuario.p_apellido}".strip() if t.instructor_origen else "—"
        instructor_destino = f"{t.instructor_destino.usuario_id_usuario.p_nombre} {t.instructor_destino.usuario_id_usuario.p_apellido}".strip() if t.instructor_destino else "—"
        ws.append([
            t.id_traslado,
            t.recurso.nombre_recurso if t.recurso_id else "",
            t.recurso.serial_recurso if t.recurso_id else "",
            t.ambiente_origen.num_ambiente if t.ambiente_origen_id else "",
            destinos_map.get(t.ambiente_destino, t.ambiente_destino) or "",
            t.fecha_traslado.strftime("%Y-%m-%d %H:%M:%S") if t.fecha_traslado else "",
            instructor_origen,
            instructor_destino,
            t.tiempo_prestamo or "—",
            t.observacion or "",
        ])

    _estilizar_excel_guarda(ws)
    return _excel_response(f"traslados_{_filters_to_suffix(filtros)}.xlsx", ws)


@never_cache
def exportar_traslados_pdf(request):
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import Table

    filtros = _get_traslados_filters(request)
    traslados = (
        TrasladoRecurso.objects.select_related("recurso", "ambiente_origen", "instructor_origen__usuario_id_usuario", "instructor_destino__usuario_id_usuario")
        .all()
        .order_by("-fecha_traslado")
    )
    traslados = _aplicar_filtros_traslados(traslados, filtros)

    destinos_ids = {t.ambiente_destino for t in traslados if t.ambiente_destino}
    destinos_map = {a.id_ambiente: a.num_ambiente for a in Ambiente.objects.filter(id_ambiente__in=destinos_ids)}

    headers = ["ID", "Recurso", "Serial", "Origen", "Destino", "Fecha", "Instructor Presta", "Instructor Recibe", "Duracion", "Observacion"]
    rows = []
    for t in traslados:
        instructor_origen = f"{t.instructor_origen.usuario_id_usuario.p_nombre} {t.instructor_origen.usuario_id_usuario.p_apellido}".strip() if t.instructor_origen else "—"
        instructor_destino = f"{t.instructor_destino.usuario_id_usuario.p_nombre} {t.instructor_destino.usuario_id_usuario.p_apellido}".strip() if t.instructor_destino else "—"
        rows.append([
            str(t.id_traslado),
            t.recurso.nombre_recurso if t.recurso_id else "",
            t.recurso.serial_recurso if t.recurso_id else "",
            str(t.ambiente_origen.num_ambiente) if t.ambiente_origen_id else "",
            str(destinos_map.get(t.ambiente_destino, t.ambiente_destino) or ""),
            t.fecha_traslado.strftime("%Y-%m-%d %H:%M:%S") if t.fecha_traslado else "",
            instructor_origen,
            instructor_destino,
            t.tiempo_prestamo or "—",
            (t.observacion or "")[:150],
        ])

    buf, doc, styles, elements = _pdf_doc("Reporte de Traslados", filtros, landscape_mode=True)
    data = _build_wrapped_pdf_data(headers, rows, styles, colors)
    table = Table(data, repeatRows=1, colWidths=[1.0 * cm, 3.0 * cm, 2.5 * cm, 1.5 * cm, 1.5 * cm, 2.5 * cm, 3.5 * cm, 3.5 * cm, 2.0 * cm, 5.0 * cm])
    table.setStyle(_pdf_table_styles(colors))
    elements.append(table)
    doc.build(elements)

    resp = HttpResponse(buf.getvalue(), content_type="application/pdf")
    resp["Content-Disposition"] = f'attachment; filename="traslados_{_filters_to_suffix(filtros)}.pdf"'
    return _no_cache(resp)


@never_cache
def exportar_ambientes_excel(request):
    filtros = _get_ambientes_filters(request)
    ambientes = _aplicar_filtros_ambientes(Ambiente.objects.all(), filtros).order_by("id_ambiente")

    wb = Workbook()
    ws = wb.active
    ws.title = "Ambientes"
    headers = ["ID", "Numero", "Capacidad", "Tipo", "Estado"]
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col_idx)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for a in ambientes:
        ws.append([a.id_ambiente, a.num_ambiente, a.capacidad, a.tipo_ambiente or "", a.estado or ""])

    _estilizar_excel_guarda(ws)
    return _excel_response(f"ambientes_{_filters_to_suffix(filtros)}.xlsx", ws)


@never_cache
def exportar_ambientes_pdf(request):
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import Table

    filtros = _get_ambientes_filters(request)
    ambientes = _aplicar_filtros_ambientes(Ambiente.objects.all(), filtros).order_by("id_ambiente")

    headers = ["ID", "Numero", "Capacidad", "Tipo", "Estado"]
    rows = []
    for a in ambientes:
        rows.append([str(a.id_ambiente), str(a.num_ambiente), str(a.capacidad), a.tipo_ambiente or "", a.estado or ""])

    buf, doc, styles, elements = _pdf_doc("Reporte de Ambientes", filtros, landscape_mode=False)
    data = _build_wrapped_pdf_data(headers, rows, styles, colors)
    table = Table(data, repeatRows=1, colWidths=[1.4 * cm, 2.2 * cm, 2.0 * cm, 3.8 * cm, 2.5 * cm])
    table.setStyle(_pdf_table_styles(colors))
    elements.append(table)
    doc.build(elements)

    resp = HttpResponse(buf.getvalue(), content_type="application/pdf")
    resp["Content-Disposition"] = f'attachment; filename="ambientes_{_filters_to_suffix(filtros)}.pdf"'
    return _no_cache(resp)
